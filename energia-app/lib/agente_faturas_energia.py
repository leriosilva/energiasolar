"""
Agente de Extração de Faturas de Energia Elétrica  v3
======================================================
Lê todos os PDFs de uma pasta, extrai os dados das faturas
e gera um arquivo Excel formatado com uma linha por fatura.

Instalação das dependências:
    pip install pdfplumber openpyxl

Uso:
    python agente_faturas_energia.py
    python agente_faturas_energia.py --pasta ./faturas --saida resultado.xlsx
    python agente_faturas_energia.py --pasta ./faturas --saida resultado.xlsx --verbose
"""

import os, re, sys, argparse, traceback, json
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print("ERRO: pdfplumber não instalado. Execute: pip install pdfplumber")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────
#  APELIDOS  –  edite conforme suas instalações
# ─────────────────────────────────────────────────────────────
APELIDOS = {
    "3071452":   "Rural Monjolinho - Itapeva MG",
    "69589968":  "R. Amajouvas 26 - São Paulo SP",
    "66447232":  "",      # ← preencha o apelido desta instalação
    "69572941":  "",      # ← preencha o apelido desta instalação
    "69572950":  "",      # ← preencha o apelido desta instalação
    "121033228": "",      # ← preencha o apelido desta instalação
}


# ─────────────────────────────────────────────────────────────
#  UTILITÁRIOS
# ─────────────────────────────────────────────────────────────

def extrair_texto(pdf_path: Path, verbose: bool = False) -> str:
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            return "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception as e:
        if verbose:
            print(f"    [AVISO] Falha ao ler {pdf_path.name}: {e}")
        return ""

def _num(s) -> float | None:
    if s is None: return None
    t = re.sub(r"[R$\s]", "", str(s)).replace(".", "").replace(",", ".")
    try: return float(t)
    except ValueError: return None

MESES_PT = {"JAN":"01","FEV":"02","MAR":"03","ABR":"04","MAI":"05","JUN":"06",
            "JUL":"07","AGO":"08","SET":"09","OUT":"10","NOV":"11","DEZ":"12"}

def _norm_mes(raw: str | None) -> str | None:
    if not raw: return None
    raw = raw.strip().upper()
    m = re.match(r"([A-Z]{3})[/\s](\d{2,4})", raw)
    if m:
        mes = MESES_PT.get(m.group(1))
        ano = m.group(2) if len(m.group(2))==4 else "20"+m.group(2)
        return f"{mes}/{ano}" if mes else None
    m = re.match(r"(\d{2})[/\-](\d{2,4})", raw)
    if m:
        ano = m.group(2) if len(m.group(2))==4 else "20"+m.group(2)
        return f"{m.group(1)}/{ano}"
    return raw

def _norm_data(raw: str | None) -> str | None:
    if not raw: return None
    raw = raw.strip()
    m = re.match(r"(\d{1,2})\s+([A-Z]{3})\s*(\d{4})?", raw.upper())
    if m:
        dia  = m.group(1).zfill(2)
        mes  = MESES_PT.get(m.group(2), "??")
        return f"{dia}/{mes}/{m.group(3)}" if m.group(3) else f"{dia}/{mes}"
    m = re.match(r"(\d{2})[/\-](\d{2})[/\-](\d{2,4})", raw)
    if m:
        ano = m.group(3) if len(m.group(3))==4 else "20"+m.group(3)
        return f"{m.group(1)}/{m.group(2)}/{ano}"
    return raw

def _bandeira(texto: str) -> str | None:
    """Retorna a bandeira do mês atual (primeira mencionada na linha de bandeira)."""
    m = re.search(
        r"Bandeira\(?s?\)?\s*tarifária\(?s?\)?[^:]*:\s*"
        r"(VERDE|AMARELA|VERMELHA\s*PATAMAR\s*[I1-9]+|VERMELHA|ESCASSEZ\s*H[IÍ]DRICA)",
        texto, re.IGNORECASE)
    if m:
        return re.sub(r"\s+", " ", m.group(1).strip()).title()
    # Fallback por precedência
    for b in ["ESCASSEZ HÍDRICA","VERMELHA PATAMAR II","VERMELHA PATAMAR I",
              "VERMELHA PATAMAR 2","VERMELHA PATAMAR 1","VERMELHA","AMARELA","VERDE"]:
        if b in texto.upper():
            return b.title()
    return None

def _mes_do_arquivo(nome: str) -> str | None:
    """Extrai MM/AAAA do nome do arquivo de forma robusta."""
    # Padrão AAAA_MM ou AAAA.MM (exige que AAAA comece com 20)
    m = re.search(r"[_.]?(20\d{2})[._](\d{2})(?:[._\s]|$)", nome)
    if m and "01" <= m.group(2) <= "12":
        return f"{m.group(2)}/{m.group(1)}"
    # Padrão inicial: 2022_05_fatura ou 2022.08.pdf
    m = re.match(r"(20\d{2})[._](\d{2})", nome)
    if m and "01" <= m.group(2) <= "12":
        return f"{m.group(2)}/{m.group(1)}"
    # Padrão Energisa: -MM-AAAA
    m = re.search(r"-(\d{2})-(20\d{2})(?:\.|$)", nome, re.IGNORECASE)
    if m and "01" <= m.group(1) <= "12":
        return f"{m.group(1)}/{m.group(2)}"
    return None

def _inst_do_arquivo(nome: str) -> str | None:
    """Extrai número de instalação do nome do arquivo."""
    # Energisa: Matricula-NNNNNNN
    m = re.search(r"Matricula[_-]0*(\d{6,10})", nome, re.IGNORECASE)
    if m: return m.group(1)
    # Enel: prefixo numérico antes de _AAAA ou .AAAA
    m = re.match(r"0*(\d{5,10})[_.](?:20\d{2})", nome)
    if m: return m.group(1)
    # Qualquer sequência de 7-10 dígitos no início
    m = re.match(r"0*(\d{7,10})", nome)
    if m: return m.group(1)
    return None


# ─────────────────────────────────────────────────────────────
#  PARSER  ENEL  SP
# ─────────────────────────────────────────────────────────────

def parse_enel(texto: str) -> dict:
    d = {}

    # ── TITULAR ──────────────────────────────────────────────
    # Layout antigo 2022: CPF/CNPJ … ISENTO  [mesX  kWhX  diasX…]  NOME
    m = re.search(
        r"CPF/CNPJ[:\s.]+[\d.*\-]+\s+INSC\.[^\n]+\n"   # linha CPF
        r"(?:[^\n]*\n){0,4}"                             # ≤4 linhas intermediárias
        r"([A-ZÁÉÍÓÚÂÊÔÀÃÕÇ]{3}[A-Za-záéíóúâêôàãõç ]{3,50})\n",  # NOME COMPLETO
        texto)
    if not m:
        # Layout novo 2023+: NOME logo após linha de datas (Monofásico …)
        m = re.search(
            r"(?:Monof[aá]sico|Trifásico|Bif[aá]sico)[^\n]*\n"
            r"([A-ZÁÉÍÓÚÂÊÔÀÃÕÇ][A-Za-záéíóúâêôàãõçÇ ]{4,55})\n",
            texto)
    if not m:
        # Nota fiscal: "PAGADOR / CPF: …  \n NOME"
        m = re.search(
            r"PAGADOR\s*/\s*CPF[^:]*:[^\n]+\n"
            r"([A-ZÁÉÍÓÚÂÊÔÀÃÕÇ][A-Za-záéíóúâêôàãõç ]{4,55})\n",
            texto)
    if m:
        nome = m.group(1).strip()
        if not re.search(r"\d|^\s*(?:R|AV|RUA|ALAMEDA)\s+|CEP|VERDE|AMARELA|VERMELHA|HÍDRICA",
                         nome.upper()):
            d["titular"] = nome.title()

    # ── INSTALAÇÃO ───────────────────────────────────────────
    m = re.search(r"(?:N[°º]\s*DA\s*INSTALA[ÇC][ÃA]O|INSTALA[ÇC][ÃA]O\s*N[°º]?)[:\s]*(\d{6,12})", texto)
    if not m:
        m = re.search(r"^(\d{7,11})\s*/", texto, re.MULTILINE)          # "0069589968 / 10007…"
    if not m:
        m = re.search(r"^(\d{8})\s+\d{8}\s+\d{2}\s+[A-Z]{3}", texto, re.MULTILINE)  # antigo
    if m:
        d["instalacao"] = m.group(1).lstrip("0") or m.group(1)

    # ── MÊS DE REFERÊNCIA ────────────────────────────────────
    # Novo layout: linha standalone "MM/AAAA" logo após o titular
    m = re.search(
        r"(?:Monof[aá]sico|Trifásico|Bif[aá]sico)[^\n]*\n[^\n]+\n((?:0[1-9]|1[0-2])/20\d{2})\n",
        texto)
    if not m:
        m = re.search(r"CONTA\s+REFERENTE\s+A[:\s]+([A-Z]{3}\s+\d{4}|\d{2}/\d{4})", texto)
    if not m:
        m = re.search(r"REFER[EÊ]NCIA[:\s]+((?:0[1-9]|1[0-2])/20\d{2})", texto)
    if m:
        d["mes"] = _norm_mes(m.group(1))

    # ── LEITURA ATUAL  &  NÚMERO DE DIAS ─────────────────────
    # Novo layout: "Monofásico  DD/MM/AAAA  DD/MM/AAAA  NN  DD/MM/AAAA"
    m = re.search(
        r"(?:Monof[aá]sico|Trifásico|Bif[aá]sico)\s+"
        r"\d{2}/\d{2}/\d{4}\s+"        # leitura anterior
        r"(\d{2}/\d{2}/\d{4})\s+"      # leitura atual
        r"(\d{2,3})\s+"                 # num dias
        r"\d{2}/\d{2}/\d{4}",           # próxima leitura
        texto)
    if m:
        d["leitura_atual"] = m.group(1)
        try: d["num_dias"] = int(m.group(2))
        except ValueError: pass
    else:
        # Layout antigo: "Leitura atual  27 MAI"
        m2 = re.search(r"Leitura\s+atual\s+(\d{2}\s+[A-Z]{3}(?:\s+\d{4})?|\d{2}/\d{2}/\d{4})",
                       texto, re.IGNORECASE)
        if m2: d["leitura_atual"] = _norm_data(m2.group(1))
        m3 = re.search(r"N[uú]mero\s+de\s+dias\s+(\d{2,3})", texto, re.IGNORECASE)
        if m3:
            try: d["num_dias"] = int(m3.group(1))
            except ValueError: pass

    # ── VENCIMENTO ───────────────────────────────────────────
    # Novo layout: "emissao  mes_ref  vencimento" na linha do boleto (pág 2)
    m = re.search(r"\d{2}/\d{2}/\d{4}\s+(?:0[1-9]|1[0-2])/20\d{2}\s+(\d{2}/\d{2}/\d{4})", texto)
    if not m:
        # Novo layout: "NF_NUM  mes_ref  vencimento  R$" linha da nota fiscal
        m = re.search(r"\d{6,12}\s+(?:0[1-9]|1[0-2])/20\d{2}\s+(\d{2}/\d{2}/\d{4})\s+R\$", texto)
    if not m:
        # Novo layout: "DD/MM/AAAA Chave de acesso" = vencimento
        m = re.search(r"(\d{2}/\d{2}/\d{4})\s+Chave\s+de\s+acesso", texto, re.IGNORECASE)
    if not m:
        # Layout antigo: "instalacao  cliente_num  DD MES AAAA  valor"
        m = re.search(r"\d{8}\s+\d{8}\s+(\d{2}\s+[A-Z]{3}\s+\d{4})\s+[\d.,]+", texto)
    if not m:
        # Qualquer "VENCIMENTO  DD/MM/AAAA" (slip / page 2)
        m = re.search(r"VENCIMENTO\s*\n\s*(\d{2}/\d{2}/\d{4})", texto)
    if m:
        d["vencimento"] = _norm_data(m.group(1))

    # ── CONSUMO (kWh) ────────────────────────────────────────
    m = re.search(r"Consumo\s+do\s+m[eê]s\s+\(kWh\)\s+([\d.,]+)", texto, re.IGNORECASE)
    if not m:
        # Novo layout: primeira linha do histórico "FEV/25  373,000  30  LID"
        m = re.search(r"[A-Z]{3}/\d{2}\s+([\d.,]+)\s+\d{2,3}\s+LID", texto)
    if not m:
        # Tabela itens: "USO SIST. DISTR. (TUSD)  KWH  373,000  ..."
        m = re.search(r"USO\s+SIST\.\s+DISTR\.[^K]+KWH\s+([\d.,]+)", texto, re.IGNORECASE)
    if not m:
        # Layout antigo: "0605 USO SIST. DISTR. (TUSD) 472,000 tarifa ..."
        m = re.search(r"0605\s+USO\s+SIST[^0-9]+([1-9][\d.,]{2,})\s+[\d.,]+\s+[\d.,]+\s+\d+%", texto)
    if not m:
        # Layout antigo: valor isolado "472,0" seguido do num_dias
        m = re.search(r"^([\d]{2,4}[,.][\d]+)\s*\n\d{2}\n[A-Z]", texto, re.MULTILINE)
    if m:
        d["consumo_kwh"] = _num(m.group(1))

    # ── TOTAL A PAGAR ────────────────────────────────────────
    # Novo layout tabela: "TOTAL  65,34  1,01  188,85  33,99"
    m = re.search(r"^TOTAL\s+([\d.,]+)\s+[\d.,]+\s+[\d.,]+", texto, re.MULTILINE)
    if not m:
        # Linha "R$65,34 Data de apresentação"
        m = re.search(r"R\$\s*([\d.,]+)\s+Data\s+de\s+apresenta", texto, re.IGNORECASE)
    if not m:
        # Linha do boleto: "emissao  nf  mes  venc  R$valor"
        m = re.search(r"\d{2}/\d{2}/\d{4}\s+\d{6,}\s+\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}\s+R\$([\d.,]+)", texto)
    if not m:
        # Layout antigo: "instalacao  cliente  DD MES AAAA  valor"
        m = re.search(r"\d{8}\s+\d{8}\s+\d{2}\s+[A-Z]{3}\s+\d{4}\s+([\d.,]+)", texto)
    if not m:
        m = re.search(r"TOTAL\s+A\s+PAGAR\s*\(?R\$?\)?\s*([\d.,]+)", texto, re.IGNORECASE)
    if m:
        d["total_pagar"] = _num(m.group(1))

    # ── BANDEIRA ─────────────────────────────────────────────
    d["bandeira"] = _bandeira(texto)

    # ── SALDO  &  ENERGIA INJETADA ───────────────────────────
    m = re.search(r"Saldo\s+atualizado\s+de\s+energia[^:]*:\s*Ativa[:\s]+([\d.,]+)", texto, re.IGNORECASE)
    if m: d["saldo_energia_kwh"] = _num(m.group(1))

    m = re.search(r"Energia\s+injetada\s+no\s+m[eê]s[^:]*:\s*Ativa[:\s]+([\d.,]+)", texto, re.IGNORECASE)
    if m: d["energia_injetada_kwh"] = _num(m.group(1))

    return d


# ─────────────────────────────────────────────────────────────
#  PARSER  ENERGISA  SS
# ─────────────────────────────────────────────────────────────

def parse_energisa(texto: str, nome_arquivo: str) -> dict:
    d = {}

    # ── TITULAR ──────────────────────────────────────────────
    m = re.search(r"(Elcio\s+Pereira\s+da\s+Silva)", texto, re.IGNORECASE)
    if not m:
        m = re.search(r"(?:DOM\.\s*BANC\..*?\n)([A-ZÁÉÍÓÚÂÊÔÀÃÕÇ][A-Za-záéíóúâêôàãõç\s]{4,55})\n", texto)
    if m:
        d["titular"] = m.group(1).strip().title()

    # ── INSTALAÇÃO ───────────────────────────────────────────
    m = re.search(r"Matricula[_-]0*(\d{6,10})", nome_arquivo, re.IGNORECASE)
    if m:
        d["instalacao"] = m.group(1).lstrip("0") or m.group(1)
    else:
        m = re.search(r"N[uú]mero\s+da\s+UC[:\s]*([\d.]+)", texto, re.IGNORECASE)
        if m: d["instalacao"] = re.sub(r"\D", "", m.group(1)).lstrip("0")

    # ── MÊS ──────────────────────────────────────────────────
    # Extraído do nome do arquivo (mais confiável)
    # (já feito em detectar_e_parsear via _mes_do_arquivo)

    # ── LEITURA ATUAL  &  DIAS ───────────────────────────────
    # "Leitura Anterior:20/04/2026 Leitura Atual:21/05/2026 Dias:31"
    m = re.search(
        r"Leitura\s+Anterior[:\s]*\d{2}/\d{2}/\d{4}\s+"
        r"Leitura\s+Atual[:\s]*(\d{2}/\d{2}/\d{4})\s+"
        r"Dias[:\s]*(\d{2,3})",
        texto, re.IGNORECASE)
    if not m:
        # Cabeçalho: "DD/MM/AAAA  DD/MM/AAAA  NN  DD/MM/AAAA"
        m = re.search(
            r"\d{2}/\d{2}/\d{4}\s+(\d{2}/\d{2}/\d{4})\s+(\d{2,3})\s+\d{2}/\d{2}/\d{4}",
            texto)
    if m:
        d["leitura_atual"] = m.group(1)
        try: d["num_dias"] = int(m.group(2))
        except ValueError: pass

    # ── VENCIMENTO ───────────────────────────────────────────
    # "MATRICULA  VENCIMENTO  Nº FATURA  TOTAL A PAGAR" section
    # "3071452-2026-5-3  04/06/2026  1839222  R$ 96,79"
    m = re.search(r"\d{5,10}[-\d]*\s+(\d{2}/\d{2}/\d{4})\s+\d{6,}", texto)
    if not m:
        # "Maio / 2026  04/06/2026  R$ 96,79"
        m = re.search(
            r"(?:[A-Z][a-z]+\s*/\s*\d{4}|(?:0[1-9]|1[0-2])/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+R\$",
            texto)
    if not m:
        m = re.search(r"VENCIMENTO\s*\n?\s*(\d{2}/\d{2}/\d{4})", texto)
    if m:
        d["vencimento"] = m.group(1)

    # ── CONSUMO ──────────────────────────────────────────────
    # "Consumo em kWh  KWH  442,00  0,951440  420,54 ..."
    m = re.search(r"Consumo\s+em\s+kWh\s+KWH\s+([\d.,]+)", texto, re.IGNORECASE)
    if not m:
        # Tabela simplificada: "KWH  442,00  tarifa  valor ..."
        m = re.search(r"^KWH\s+([\d.,]+)\s+[\d.,]+\s+[\d.,]+", texto, re.MULTILINE)
    if not m:
        # Tabela detalhada: "Energia ativa em kWh  Ponta  NNN  NNN  1  442"
        m = re.search(r"Energia\s+ativa\s+em\s+kWh\s+Ponta\s+\d+\s+\d+\s+1\s+([\d.,]+)", texto, re.IGNORECASE)
    if not m:
        # Estrutura do consumo: "KWH Ponta  15042  14600  1  0  0  0  442  442"
        m = re.search(r"KWH\s+Ponta\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+([\d.,]+)\s+\1", texto)
    if not m:
        # Último fallback: "FATURADO\n...\n442"  via histórico
        m = re.search(r"(?:CONSUMO\s+FATURADO|Consumo\s+Faturado)[^\n]*\n[^\n]*?\b([\d]+[.,][\d]{3})\b", texto, re.IGNORECASE)
    if m:
        d["consumo_kwh"] = _num(m.group(1))

    # ── TOTAL ────────────────────────────────────────────────
    # "TOTAL:  96,79  3,50  96,79  17,42"
    m = re.search(r"^TOTAL[:\s]+([\d.,]+)\s+[\d.,]+\s+[\d.,]+", texto, re.MULTILINE)
    if not m:
        m = re.search(r"TOTAL\s+A\s+PAGAR[^\d]*([\d.,]+)", texto, re.IGNORECASE)
    if m:
        d["total_pagar"] = _num(m.group(1))

    # ── BANDEIRA ─────────────────────────────────────────────
    d["bandeira"] = _bandeira(texto)

    # ── SALDO  &  INJETADA ───────────────────────────────────
    m = re.search(r"Saldo\s+Acumulado[:\s]*([\d.,]+)", texto, re.IGNORECASE)
    if m: d["saldo_energia_kwh"] = _num(m.group(1))

    m = re.search(r"INJ\s+Ponta\s+[\d.]+\s+[\d.]+\s+1\s+([\d.,]+)", texto, re.IGNORECASE)
    if not m:
        m = re.search(r"Energia\s+injetada\s+Ponta\s+\d+\s+\d+\s+1\s+([\d.,]+)", texto, re.IGNORECASE)
    if m: d["energia_injetada_kwh"] = _num(m.group(1))

    return d


# ─────────────────────────────────────────────────────────────
#  DETECÇÃO  &  ROTEAMENTO
# ─────────────────────────────────────────────────────────────

def detectar_e_parsear(texto: str, nome_arquivo: str) -> dict:
    nome_up = nome_arquivo.upper()

    if "ENERGISA" in nome_up:
        d = parse_energisa(texto, nome_arquivo)
        d.setdefault("distribuidora", "Energisa Sul-Sudeste")
    else:
        d = parse_enel(texto)
        d.setdefault("distribuidora", "Enel Distribuição SP")

    # Instalação fallback
    if not d.get("instalacao"):
        inst = _inst_do_arquivo(nome_arquivo)
        if inst: d["instalacao"] = inst

    # Normalizar instalação (remover zeros à esquerda para lookup)
    inst_raw = d.get("instalacao", "")
    inst_norm = inst_raw.lstrip("0") or inst_raw

    # Mês fallback
    if not d.get("mes"):
        d["mes"] = _mes_do_arquivo(nome_arquivo)

    # Apelido (testa com e sem zeros à esquerda)
    d["apelido"] = (APELIDOS.get(inst_norm)
                    or APELIDOS.get(inst_raw)
                    or "")

    d.setdefault("titular", "")
    return d


# ─────────────────────────────────────────────────────────────
#  GERAÇÃO  DO  EXCEL
# ─────────────────────────────────────────────────────────────

COLUNAS = [
    ("Distribuidora",              18),
    ("Titular",                    28),
    ("Instalação",                 14),
    ("Apelido",                    24),
    ("Mês",                        10),
    ("Leitura Atual",              14),
    ("Vencimento",                 14),
    ("Consumo Faturado (kWh)",     22),
    ("# Dias",                      8),
    ("Total a Pagar (R$)",         16),
    ("Bandeira",                   24),
    ("Saldo de Energia (kWh)",     22),
    ("Energia Injetada (kWh)",     22),
    ("Arquivo",                    38),
    ("Status",                      8),
]

CAMPO_COL = {
    "distribuidora":"Distribuidora", "titular":"Titular",
    "instalacao":"Instalação",       "apelido":"Apelido",
    "mes":"Mês",                     "leitura_atual":"Leitura Atual",
    "vencimento":"Vencimento",       "consumo_kwh":"Consumo Faturado (kWh)",
    "num_dias":"# Dias",             "total_pagar":"Total a Pagar (R$)",
    "bandeira":"Bandeira",           "saldo_energia_kwh":"Saldo de Energia (kWh)",
    "energia_injetada_kwh":"Energia Injetada (kWh)",
}

BAND_COR = {
    "Verde":              ("00B050","FFFFFF"),
    "Amarela":            ("FFC000","7B4700"),
    "Vermelha":           ("FF0000","FFFFFF"),
    "Vermelha Patamar I": ("FF4444","FFFFFF"),
    "Vermelha Patamar 1": ("FF4444","FFFFFF"),
    "Vermelha Patamar Ii":("CC0000","FFFFFF"),
    "Vermelha Patamar 2": ("CC0000","FFFFFF"),
    "Escassez Hídrica":   ("FF6600","FFFFFF"),
}

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def gerar_excel(registros: list, caminho: str):
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Faturas"
    ci  = {nome: i+1 for i,(nome,_) in enumerate(COLUNAS)}
    hf  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hft = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    bd  = _border()

    for i,(nome,larg) in enumerate(COLUNAS, 1):
        c = ws.cell(row=1, column=i, value=nome)
        c.fill=hf; c.font=hft; c.border=bd
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}1"

    bf = Font(name="Arial", size=10)

    for ri, reg in enumerate(registros, 2):
        alt  = ri % 2 == 0
        rf   = PatternFill("solid",
                           start_color="D6E4F0" if alt else "FFFFFF",
                           end_color  ="D6E4F0" if alt else "FFFFFF")

        def w(col, val, fmt=None, align="left"):
            if col not in ci or val is None: return
            c = ws.cell(row=ri, column=ci[col], value=val)
            c.fill=rf; c.font=bf; c.border=bd
            c.alignment = Alignment(horizontal=align, vertical="center")
            if fmt: c.number_format = fmt

        for campo, col in CAMPO_COL.items():
            if campo == "bandeira": continue
            v = reg.get(campo)
            if v is None: continue
            if campo == "consumo_kwh":    w(col, float(v), "#,##0.000", "right")
            elif campo == "total_pagar":  w(col, float(v), '"R$" #,##0.00', "right")
            elif campo in ("saldo_energia_kwh","energia_injetada_kwh"):
                                          w(col, float(v), "#,##0.0", "right")
            elif campo == "num_dias":     w(col, int(v), "0", "center")
            elif campo in ("mes","leitura_atual","vencimento","instalacao"):
                                          w(col, str(v), align="center")
            else:                         w(col, v)

        # Bandeira colorida
        band = reg.get("bandeira")
        if band:
            c = ws.cell(row=ri, column=ci["Bandeira"], value=band)
            cor = BAND_COR.get(band)
            if cor:
                bg, fg = cor
                c.fill = PatternFill("solid", start_color=bg, end_color=bg)
                c.font = Font(name="Arial", size=10, bold=True, color=fg)
            else:
                c.fill=rf; c.font=bf
            c.border=bd
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Arquivo
        c = ws.cell(row=ri, column=ci["Arquivo"], value=reg.get("arquivo",""))
        c.fill=rf; c.font=Font(name="Arial",size=9,color="5F5E5A"); c.border=bd
        c.alignment = Alignment(vertical="center")

        # Status
        ok = reg.get("_ok", True)
        c  = ws.cell(row=ri, column=ci["Status"], value="OK" if ok else "ERRO")
        c.fill  = PatternFill("solid",
                              start_color="E2EFDA" if ok else "FCEBEB",
                              end_color  ="E2EFDA" if ok else "FCEBEB")
        c.font  = Font(name="Arial", size=10, bold=True,
                       color="375623" if ok else "A32D2D")
        c.border=bd
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Linha de totais
    n  = len(registros)
    tr = n + 2
    tf = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    tt = Font(bold=True, name="Arial", size=10, color="FFFFFF")
    for i in range(1, len(COLUNAS)+1):
        c = ws.cell(row=tr, column=i); c.fill=tf; c.border=bd
    ws.cell(row=tr, column=1, value="TOTAIS").font = tt
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for col, fmt, align in [
        ("Consumo Faturado (kWh)", "#,##0.000",     "right"),
        ("Total a Pagar (R$)",     '"R$" #,##0.00', "right"),
        ("# Dias",                 "0",              "center"),
    ]:
        c  = ws.cell(row=tr, column=ci[col],
                     value=f"=SUM({get_column_letter(ci[col])}2:{get_column_letter(ci[col])}{n+1})")
        c.number_format=fmt; c.font=tt
        c.alignment=Alignment(horizontal=align, vertical="center")

    # Aba resumo por instalação
    from collections import defaultdict
    ws2 = wb.create_sheet("Resumo por Instalação")
    h2  = ["Instalação","Apelido","Qtd Faturas","Consumo Total (kWh)","Gasto Total (R$)","Distribuidora"]
    for i,h in enumerate(h2,1):
        c=ws2.cell(row=1,column=i,value=h)
        c.fill=hf; c.font=hft; c.border=bd
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for i,w_ in enumerate([14,28,12,22,16,22],1):
        ws2.column_dimensions[get_column_letter(i)].width=w_

    res = defaultdict(lambda:{"ap":"","n":0,"kwh":0.0,"tot":0.0,"dist":""})
    for r in registros:
        k = r.get("instalacao","?")
        res[k]["ap"]   = r.get("apelido") or res[k]["ap"]
        res[k]["n"]   += 1
        res[k]["kwh"] += r.get("consumo_kwh") or 0
        res[k]["tot"] += r.get("total_pagar") or 0
        res[k]["dist"] = r.get("distribuidora","")

    for ri2,(inst,v) in enumerate(sorted(res.items()),2):
        alt2 = ri2%2==0
        rf2  = PatternFill("solid",
                           start_color="D6E4F0" if alt2 else "FFFFFF",
                           end_color  ="D6E4F0" if alt2 else "FFFFFF")
        for ci2,val in enumerate([inst,v["ap"],v["n"],
                                   round(v["kwh"],3),round(v["tot"],2),v["dist"]],1):
            c=ws2.cell(row=ri2,column=ci2,value=val)
            c.fill=rf2; c.font=bf; c.border=bd
            if ci2==4: c.number_format="#,##0.000"; c.alignment=Alignment(horizontal="right",vertical="center")
            elif ci2==5: c.number_format='"R$" #,##0.00'; c.alignment=Alignment(horizontal="right",vertical="center")
            elif ci2==3: c.alignment=Alignment(horizontal="center",vertical="center")
            else: c.alignment=Alignment(vertical="center")

    wb.save(caminho)


# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────

def _coletar_pdfs(paths):
    """Recebe lista de arquivos/pastas e devolve lista de PDFs unicos."""
    vistos, pdfs = set(), []
    for raw in paths:
        p = Path(raw)
        candidatos = []
        if p.is_dir():
            candidatos = [x for x in sorted(p.rglob("*")) if x.suffix.lower() == ".pdf"]
        elif p.suffix.lower() == ".pdf":
            candidatos = [p]
        for c in candidatos:
            try:
                if c.stat().st_size == 0:
                    continue
            except OSError:
                continue
            chave = str(c).lower()
            if chave not in vistos:
                vistos.add(chave); pdfs.append(c)
    return pdfs


def extrair_para_json(paths, verbose=False):
    """Extrai os dados de cada PDF e devolve uma lista de dicts (JSON-serializavel)."""
    registros = []
    for pdf in _coletar_pdfs(paths):
        try:
            texto = extrair_texto(pdf, verbose)
            if not texto.strip():
                registros.append({"arquivo": pdf.name, "_ok": False, "instalacao": "", "mes": "", "_motivo": "sem texto extraivel"})
                continue
            d = detectar_e_parsear(texto, pdf.name)
            d["arquivo"] = pdf.name
            d["_ok"] = True
            registros.append(d)
        except Exception as e:
            registros.append({"arquivo": pdf.name, "_ok": False, "instalacao": "", "mes": "", "_motivo": f"erro: {e}"})
    return registros


def main():
    ap = argparse.ArgumentParser(description="Extrai dados de faturas de energia e gera Excel ou JSON.")
    ap.add_argument("--pasta",   default=".", help="Pasta com os PDFs (padrão: pasta atual)")
    ap.add_argument("--saida",   default="faturas_energia.xlsx", help="Arquivo Excel de saída")
    ap.add_argument("--json",    action="store_true", help="Emite JSON no stdout (uma fatura por PDF) em vez de gerar Excel")
    ap.add_argument("--verbose", action="store_true")
    ap.add_argument("arquivos",  nargs="*", help="PDFs ou pastas específicos (se vazio, usa --pasta)")
    args = ap.parse_args()

    # ── MODO JSON (integracao com o backend) ─────────────────
    if args.json:
        paths = args.arquivos if args.arquivos else [args.pasta]
        registros = extrair_para_json(paths, args.verbose)
        sys.stdout.write(json.dumps(registros, ensure_ascii=False))
        sys.stdout.flush()
        return

    pasta = Path(args.pasta).resolve()
    if not pasta.exists():
        print(f"ERRO: pasta não encontrada: {pasta}"); sys.exit(1)

    # Coleta PDFs sem duplicatas (Windows: rglob case-insensitive)
    vistos, pdfs = set(), []
    for p in sorted(pasta.rglob("*")):
        if p.suffix.lower() == ".pdf" and p.stat().st_size > 0:
            chave = str(p).lower()
            if chave not in vistos:
                vistos.add(chave); pdfs.append(p)

    if not pdfs:
        print(f"Nenhum PDF encontrado em: {pasta}"); sys.exit(0)

    print(f"\n{'='*62}\n  Agente de Faturas de Energia Elétrica  v3\n{'='*62}")
    print(f"  Pasta  : {pasta}")
    print(f"  PDFs   : {len(pdfs)}")
    print(f"  Saída  : {args.saida}\n{'='*62}\n")

    registros, erros = [], []

    for i, pdf in enumerate(pdfs, 1):
        pfx = f"  [{i:3d}/{len(pdfs)}]"
        try:
            texto = extrair_texto(pdf, args.verbose)
            if not texto.strip():
                print(f"{pfx} [SEM TEXTO] {pdf.name}")
                erros.append(pdf.name)
                registros.append({"arquivo":pdf.name,"_ok":False,
                                   "instalacao":"","mes":"","distribuidora":"Sem texto"})
                continue

            d = detectar_e_parsear(texto, pdf.name)
            d["arquivo"] = pdf.name
            d["_ok"]     = True
            registros.append(d)

            inst  = d.get("instalacao","?")
            mes   = d.get("mes","?")
            total = d.get("total_pagar")
            ts    = f"R$ {total:,.2f}" if total is not None else "R$ ?"
            print(f"{pfx} {inst:>12} | {mes:>7} | {ts:>12}  {pdf.name}")

            if args.verbose:
                for k,v in d.items():
                    if not k.startswith("_") and k != "arquivo":
                        print(f"{'':>22}{k}: {v}")

        except Exception as e:
            print(f"{pfx} [ERRO] {pdf.name}: {e}")
            if args.verbose: traceback.print_exc()
            erros.append(pdf.name)
            registros.append({"arquivo":pdf.name,"_ok":False,
                               "instalacao":"","mes":"","distribuidora":f"Erro: {e}"})

    print(f"\n{'='*62}")
    print(f"  Processados : {len(registros)}")
    print(f"  Com erro    : {len(erros)}")
    gerar_excel(registros, args.saida)
    print(f"  Excel salvo : {args.saida}")
    print(f"  Total pago  : R$ {sum(r.get('total_pagar') or 0 for r in registros):,.2f}")
    print(f"  Total kWh   : {sum(r.get('consumo_kwh') or 0 for r in registros):,.0f} kWh")
    print(f"{'='*62}\n")

    if erros:
        print("Arquivos com problema:")
        for e in erros: print(f"  - {e}")


if __name__ == "__main__":
    main()
